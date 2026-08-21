from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import os
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pdfplumber
import xlrd


# ── HCN output headers ──────────────────────────────────────────────────────
BALANCETE_HCN_HEADERS = [
    "REG", "ContaContabil", "NomeConta", "SaldoAnterior",
    "Debito", "Credito", "SaldoAtual",
    "Estoque", "AtivoImob", "DepreciacaoAcumulativa",
    "ContaFinanceira", "ReservaDeContingência", "Centro de Custos",
]
DIARIO_HCN_HEADERS = [
    "REG", "DATA", "CLASSIFICAÇÃO", "DESCRIÇÃO",
    "HISTÓRICO", "DÉBITO", "CRÉDITO", "Centro de Custos", "Mov",
]
RAZAO_HCN_HEADERS = [
    "REG", "NOME CONTA", "CONTA CONTÁBIL", "SALDO ANTERIOR", "DATA",
    "HISTÓRICO", "DÉBITO", "CRÉDITO", "Saldo", "Contra Partida", "Centro de Custos",
]

DEFAULT_OUTPUT_HEADERS = ["Data", "Descricao", "Debito", "Credito", "Saldo"]

# ── Styles ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill("solid", fgColor="F4F8FB")
TOTAL_BORDER = Border(
    bottom=Side(style="medium", color="1F1F1F"),
    top=Side(style="thin", color="D9E2F3"),
)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
ACCOUNTING_FORMAT      = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* #,##0.00_-;_-@_-'
# Razão: zero exibido como traço (-)
ACCOUNTING_FORMAT_DASH = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"_-;_-@_-'
# Razão SALDO ANTERIOR: prefixo R$, zero como traço
ACCOUNTING_FORMAT_BRL  = '_-"R$ "* #,##0.00_-;\\-"R$ "* #,##0.00_-;_-"R$ "* "-"_-;_-@_-'

# ── Regexes ──────────────────────────────────────────────────────────────────
DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
MONEY_RE = re.compile(
    r"^-?\s*(?:R\$\s*)?\d{1,3}(?:\.\d{3})*,\d{2}$"
    r"|^-?\s*(?:R\$\s*)?\d+,\d{2}$"
)
ACCOUNT_RE = re.compile(r"^([\d]+)\s+([\d.]+)\s*-\s*(.+)$")
CONTRA_RE = re.compile(r"^(\d+)\s*-\s*([\d.].*)$")

# ── Portuguese month names (normalised, no accents) ──────────────────────────
PT_MONTHS: dict[str, str] = {
    "janeiro": "01", "fevereiro": "02", "marco": "03",
    "abril": "04", "maio": "05", "junho": "06", "julho": "07",
    "agosto": "08", "setembro": "09", "outubro": "10",
    "novembro": "11", "dezembro": "12",
}

HEADER_ALIASES = {
    "date": {"data", "dt"},
    "description": {"historico", "descricao", "descricao historico", "complemento", "detalhe"},
    "debit": {"debito", "valor debito"},
    "credit": {"credito", "valor credito"},
    "saldo": {"saldo", "saldo final"},
}


# ═══════════════════════════════════════════════════════════════════════════════
# Public entry point
# ═══════════════════════════════════════════════════════════════════════════════

def beautify_workbook(file_path: str, input_extension: str = ".xlsx") -> BytesIO:
    if input_extension == ".pdf":
        with open(file_path, "rb") as fh:
            return beautify_pdf(BytesIO(fh.read()))

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)

    # Accumulate rows per document type so sheets of the same type get merged.
    # Key: tuple of header names → (first_title, parsed_sheet_template, [rows])
    merged: dict[tuple, tuple[str, dict, list]] = {}
    order: list[tuple] = []  # insertion order to preserve sheet ordering

    for original_title, rows in read_input_sheets(file_path, input_extension):
        parsed_sheet = extract_records(rows, sheet_title=original_title)
        if not parsed_sheet:
            continue

        key = tuple(parsed_sheet["headers"])
        if key not in merged:
            merged[key] = (original_title, parsed_sheet, list(parsed_sheet["rows"]))
            order.append(key)
        else:
            merged[key][2].extend(parsed_sheet["rows"])

    created_sheets = 0
    for key in order:
        title, template, all_rows = merged[key]
        if not all_rows:
            continue
        final_sheet = dict(template)
        final_sheet["rows"] = all_rows

        output_sheet = output_workbook.create_sheet(
            title=build_sheet_title(final_sheet["headers"], created_sheets)
        )
        write_records(output_sheet, final_sheet)
        style_output_sheet(output_sheet, final_sheet, created_sheets)
        created_sheets += 1

    if created_sheets == 0:
        raise ValueError(
            "Nao encontrei lancamentos no formato esperado. "
            "Se quiser, me envie um exemplo desse Excel para eu ajustar o parser."
        )

    output = BytesIO()
    output_workbook.save(output)
    output.seek(0)
    return output


def beautify_pdf(file_stream: BytesIO) -> BytesIO:
    parsed_documents = parse_pdf_documents(file_stream)
    if not parsed_documents:
        raise ValueError(
            "Nao consegui interpretar esse PDF ainda. "
            "Se ele for exportado do sistema contabil, me envie o arquivo que eu ajusto o layout."
        )

    output_workbook = Workbook()
    output_workbook.remove(output_workbook.active)

    for sheet_index, (title, parsed_sheet) in enumerate(parsed_documents):
        output_sheet = output_workbook.create_sheet(
            title=build_sheet_title(parsed_sheet["headers"], sheet_index)
        )
        write_records(output_sheet, parsed_sheet)
        style_output_sheet(output_sheet, parsed_sheet, sheet_index)

    output = BytesIO()
    output_workbook.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════════════════════════
# PDF parsing
# ═══════════════════════════════════════════════════════════════════════════════

def parse_pdf_documents(file_stream: BytesIO) -> list[tuple[str, dict]]:
    file_stream.seek(0)
    with pdfplumber.open(file_stream) as pdf:
        first_page_text = pdf.pages[0].extract_text() or ""
        norm = normalize_text(first_page_text)

        if "balancete" in norm:
            rows = parse_balancete_pdf(pdf)
            if rows:
                return [("BalancetePDF", _balancete_sheet(rows))]

        if "diario geral" in norm or "diario" in norm:
            rows = parse_diario_pdf(pdf)
            if rows:
                return [("DiarioPDF", _diario_sheet(rows))]

        if "razao contabil" in norm or "razao" in norm:
            rows = parse_razao_pdf(pdf)
            if rows:
                return [("RazaoPDF", _razao_sheet(rows))]

    return []


def _balancete_sheet(rows: list[dict]) -> dict:
    return {
        "headers": BALANCETE_HCN_HEADERS,
        "rows": rows,
        "date_columns": set(),
        "money_columns": {4, 5, 6, 7},
        "description_column": 3,
    }


def _diario_sheet(rows: list[dict]) -> dict:
    return {
        "headers": DIARIO_HCN_HEADERS,
        "rows": rows,
        "date_columns": set(),
        "money_columns": {6, 7},
        "description_column": 5,
    }


def _razao_sheet(rows: list[dict]) -> dict:
    # Col positions (1-indexed): REG=1 NOME CONTA=2 CONTA=3 SALDO_ANT=4 DATA=5
    # HISTÓRICO=6 DÉBITO=7 CRÉDITO=8 Saldo=9 Contra Partida=10 Centro=11
    return {
        "headers": RAZAO_HCN_HEADERS,
        "rows": rows,
        "date_columns": set(),
        "money_columns": {4, 7, 8, 9},
        "description_column": 6,
    }


# ── PDF page helpers ──────────────────────────────────────────────────────────

def parse_balancete_pdf(pdf) -> list[dict]:
    """Parse TOTVS Balancete PDF lines.

    Line format (no Red column):
      CONTA  DESCRICAO  SALDO_ANTERIOR[D|C]  DEBITO  CREDITO  SALDO_ATUAL[D|C]
    """
    rows: list[dict] = []
    _money = r"-?\d{1,3}(?:\.\d{3})*,\d{2}"
    pattern = re.compile(
        r"^(?P<conta>\d[\d.]*)\s+"
        r"(?P<descricao>.+?)\s+"
        r"(?P<saldo_anterior>" + _money + r")[DC]?\s+"
        r"(?P<debito>" + _money + r")\s+"
        r"(?P<credito>" + _money + r")\s+"
        r"(?P<saldo_atual>" + _money + r")[DC]?$"
    )
    for page in pdf.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            m = pattern.match(normalize_spaces(line))
            if not m:
                continue
            rows.append({
                "REG": "0300",
                "ContaContabil": m.group("conta"),
                "NomeConta": m.group("descricao"),
                "SaldoAnterior": parse_money_value(m.group("saldo_anterior")),
                "Debito": parse_money_value(m.group("debito")),
                "Credito": parse_money_value(m.group("credito")),
                "SaldoAtual": parse_money_value(m.group("saldo_atual")),
                "Estoque": "N",
                "AtivoImob": "N",
                "DepreciacaoAcumulativa": "N",
                "ContaFinanceira": "N",
                "ReservaDeContingência": "N",
                "Centro de Custos": None,
            })
    return rows


def parse_diario_pdf(pdf) -> list[dict]:
    rows: list[dict] = []
    current: dict | None = None

    for page in pdf.pages:
        line_map = extract_pdf_lines(page)
        for _, words in line_map:
            if not words:
                continue
            ordered = sorted(words, key=lambda item: item[0])
            texts = [t for _, t in ordered]
            first_text = texts[0]
            line_text = " ".join(texts)
            norm_line = normalize_text(line_text)

            if (
                first_text == "Lote"
                or "diario" in norm_line
                or "pagina:" in norm_line
                or "periodo:" in norm_line
                or "uruacu" in norm_line
                or "hcn - hosp" in norm_line
                or "total do dia:" in norm_line
                or "total da empresa:" in norm_line
                or "total geral:" in norm_line
                or re.fullmatch(r"(?:-?\d{1,3}(?:\.\d{3})*,\d{2}\s*){1,3}", line_text.strip())
                or "19.324.171/0008-70" in line_text
                or "centro-norte goiano" in norm_line
            ):
                continue

            if re.fullmatch(r"\d+", first_text) and len(ordered) >= 2 and re.match(r"^\d{8}", texts[1]):
                if current is not None:
                    rows.extend(_finalise_diario_pdf(current))

                lote = first_text
                nr_mvto, conta_debito_codigo = split_mvto_and_account(texts[1])
                debit_tokens = [conta_debito_codigo] if conta_debito_codigo else []
                debit_tokens.extend(t for x, t in ordered if 100 <= x < 200)
                credit_tokens = [t for x, t in ordered if 200 <= x < 309]
                historico_words = [t for x, t in ordered if 309 <= x < 470]
                amount_values = [
                    (x, parse_money_value(t))
                    for x, t in ordered
                    if x >= 470 and parse_money_value(t) is not None
                ]

                debito = credito = None
                if len(amount_values) >= 2:
                    debito = amount_values[0][1]
                    credito = amount_values[-1][1]
                elif len(amount_values) == 1:
                    amount = amount_values[0][1]
                    if bool(debit_tokens) and not bool(credit_tokens):
                        debito = amount
                    elif bool(credit_tokens) and not bool(debit_tokens):
                        credito = amount
                    elif amount_values[0][0] >= 535:
                        credito = amount
                    else:
                        debito = amount

                current = {
                    "_lote": lote,
                    "_nr_mvto": nr_mvto,
                    "_conta_debito": " ".join(debit_tokens).strip(),
                    "_conta_credito": " ".join(credit_tokens).strip(),
                    "_historico": " ".join(historico_words).strip(),
                    "_debito": debito,
                    "_credito": credito,
                    "_debito_desc": [],
                    "_credito_desc": [],
                    "_historico_extra": [],
                }
                continue

            if current is None:
                continue
            if "emitido por:" in norm_line:
                continue

            for x, text in ordered:
                if 100 <= x < 200:
                    current["_debito_desc"].append(text)
                elif 200 <= x < 309:
                    current["_credito_desc"].append(text)
                elif 309 <= x < 470:
                    current["_historico_extra"].append(text)

    if current is not None:
        rows.extend(_finalise_diario_pdf(current))
    return rows


def _finalise_diario_pdf(c: dict) -> list[dict]:
    conta_debito = clean_diario_account(
        join_description(c["_conta_debito"], " ".join(c.pop("_debito_desc", [])))
    )
    conta_credito = clean_diario_account(
        join_description(c["_conta_credito"], " ".join(c.pop("_credito_desc", [])))
    )
    historico = join_description(
        c["_historico"], " ".join(c.pop("_historico_extra", []))
    ) or "Sem historico"

    rows: list[dict] = []
    if conta_debito:
        code, name = split_account_field(conta_debito)
        rows.append({
            "REG": 1600, "DATA": "", "CLASSIFICAÇÃO": code, "DESCRIÇÃO": name,
            "HISTÓRICO": historico, "DÉBITO": decimal_to_float(c["_debito"]),
            "CRÉDITO": 0.0 if c["_debito"] is not None else None, "Centro de Custos": None,
        })
    if conta_credito:
        code, name = split_account_field(conta_credito)
        rows.append({
            "REG": 1600, "DATA": "", "CLASSIFICAÇÃO": code, "DESCRIÇÃO": name,
            "HISTÓRICO": historico, "DÉBITO": None,
            "CRÉDITO": decimal_to_float(c["_credito"]), "Centro de Custos": None,
        })
    return rows


def parse_razao_pdf(pdf) -> list[dict]:
    rows: list[dict] = []
    current_account_name = ""
    current_account_code = ""
    current_saldo_anterior: Decimal | None = None
    current_date = ""
    pending_record: dict | None = None
    money_pattern = r"-?\d{1,3}(?:\.\d{3})*,\d{2}"

    for page in pdf.pages:
        text = page.extract_text() or ""
        for raw_line in text.splitlines():
            line = normalize_spaces(raw_line)
            if not line:
                continue
            if line.startswith("Conta Anal") and "tica:" in line:
                m = re.search(r"Conta Anal[íi]tica:\s*(.+?)\s+Saldo Anterior:", line)
                account_full = m.group(1).strip() if m else line.split(":", 1)[-1].strip()
                current_account_code, current_account_name = parse_razao_account_header(account_full)
                saldo_m = re.search(r"Saldo Anterior:\s*(" + money_pattern + r")", line)
                current_saldo_anterior = parse_money_value(saldo_m.group(1)) if saldo_m else None
                pending_record = None
                continue
            if line.startswith("Data ") or "Razão" in line or "Pagina:" in normalize_text(line):
                continue
            parsed_date = parse_date_value(line.split(" ")[0])
            if parsed_date:
                current_date = parsed_date
                values = re.findall(money_pattern, line)
                pending_record = {
                    "REG": 1700,
                    "NOME CONTA": current_account_name or "Sem conta",
                    "CONTA CONTÁBIL": current_account_code or "",
                    "SALDO ANTERIOR": decimal_to_float(current_saldo_anterior),
                    "DATA": current_date,
                    "HISTÓRICO": "Sem historico",
                    "DÉBITO": decimal_to_float(parse_money_value(values[0])) if len(values) > 0 else None,
                    "CRÉDITO": decimal_to_float(parse_money_value(values[1])) if len(values) > 1 else None,
                    "Saldo": decimal_to_float(parse_money_value(values[2])) if len(values) > 2 else None,
                    "Contra Partida": "",
                    "Centro de Custos": None,
                }
                rows.append(pending_record)
                continue
            if line.startswith("Contrapartida:"):
                if pending_record is not None:
                    contra = line.split(":", 1)[1].strip()
                    mm = re.search(rf"\s+{money_pattern}\s+{money_pattern}\s+", contra)
                    if mm:
                        contra = contra[: mm.start()].strip()
                    pending_record["Contra Partida"] = parse_razao_contra_code(contra)
                continue
            if pending_record is not None:
                extra = line.strip()
                if extra and normalize_text(pending_record["HISTÓRICO"]) in {"", "sem historico"}:
                    pending_record["HISTÓRICO"] = extra
    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# XLS / XLSX extraction
# ═══════════════════════════════════════════════════════════════════════════════

def extract_records(rows: list[tuple], sheet_title: str = "") -> dict | None:
    non_empty_rows = [list(row) for row in rows if not row_is_empty(row)]
    if not non_empty_rows:
        return None

    balancete = detect_balancete_layout(non_empty_rows)
    if balancete is not None:
        balancete_rows = parse_balancete_rows(non_empty_rows, balancete)
        if balancete_rows:
            return {
                "headers": BALANCETE_HCN_HEADERS,
                "rows": balancete_rows,
                "date_columns": set(),
                "money_columns": {4, 5, 6, 7},
                "description_column": 3,
            }

    diario = detect_diario_layout(non_empty_rows)
    if diario is not None:
        diario_rows = parse_diario_rows(non_empty_rows, diario)
        if diario_rows:
            return {
                "headers": DIARIO_HCN_HEADERS,
                "rows": diario_rows,
                "date_columns": set(),
                "money_columns": {6, 7},
                "description_column": 5,
            }

    soulmv_razao = detect_soulmv_razao_layout(non_empty_rows, sheet_title)
    if soulmv_razao is not None:
        soulmv_rows = parse_soulmv_razao_rows(non_empty_rows, soulmv_razao)
        if soulmv_rows:
            return {
                "headers": RAZAO_HCN_HEADERS,
                "rows": soulmv_rows,
                "date_columns": set(),
                "money_columns": {4, 7, 8, 9},
                "description_column": 6,
            }

    razao = detect_razao_layout(non_empty_rows)
    if razao is not None:
        razao_rows = parse_razao_rows(non_empty_rows, razao)
        if razao_rows:
            return {
                "headers": RAZAO_HCN_HEADERS,
                "rows": razao_rows,
                "date_columns": set(),
                "money_columns": {4, 7, 8, 9},
                "description_column": 6,
            }

    structured = detect_structured_layout(non_empty_rows)
    if structured is not None:
        structured_records = parse_structured_rows(
            non_empty_rows, structured["header_index"], structured["columns"]
        )
        if structured_records:
            return {
                "headers": DEFAULT_OUTPUT_HEADERS,
                "rows": structured_records,
                "date_columns": {1},
                "money_columns": {3, 4, 5},
                "description_column": 2,
            }

    generic_records = parse_generic_rows(non_empty_rows)
    if generic_records:
        return {
            "headers": DEFAULT_OUTPUT_HEADERS,
            "rows": generic_records,
            "date_columns": {1},
            "money_columns": {3, 4, 5},
            "description_column": 2,
        }
    return None


def read_input_sheets(file_path: str, input_extension: str) -> list[tuple[str, list[tuple]]]:
    if input_extension == ".pdf":
        with open(file_path, "rb") as fh:
            return read_pdf_sheets(BytesIO(fh.read()))

    if input_extension == ".xls":
        with open(file_path, "rb") as fh:
            file_contents = fh.read()
        workbook = xlrd.open_workbook(file_contents=file_contents)
        sheets: list[tuple[str, list[tuple]]] = []
        for sheet in workbook.sheets():
            rows = []
            for row_index in range(sheet.nrows):
                parsed_row = []
                for col_index in range(sheet.ncols):
                    parsed_row.append(
                        convert_xls_cell(
                            workbook,
                            sheet.cell_value(row_index, col_index),
                            sheet.cell_type(row_index, col_index),
                        )
                    )
                rows.append(tuple(parsed_row))
            sheets.append((sheet.name, rows))
        return sheets

    # Use read_only=True so openpyxl streams rows from the zip archive one at a
    # time rather than deserialising the entire workbook into memory up front.
    # keep_vba is incompatible with read_only mode, so fall back for .xlsm.
    if input_extension == ".xlsm":
        workbook = load_workbook(file_path, data_only=True, keep_vba=True)
        sheets_data = [
            (sheet.title, list(sheet.iter_rows(values_only=True)))
            for sheet in workbook.worksheets
        ]
        workbook.close()
        return sheets_data

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        file_size = os.path.getsize(file_path)
        return [
            (sheet.title, _read_worksheet_rows(sheet, file_size))
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _read_worksheet_rows(sheet, file_size: int) -> list[tuple]:
    # Some source systems (e.g. SOULMV) write a stale <dimension ref="A1"/> tag
    # even when the sheet holds thousands of real rows. openpyxl's read_only
    # mode trusts that tag at face value, so max_row/max_column report 1 and
    # iter_rows() silently yields nothing. A reported dimension that tiny on a
    # file far too large to actually be empty is the tell; force a real
    # recalculation before reading in that case.
    if sheet.max_row <= 1 and sheet.max_column <= 1 and file_size > 20_000:
        sheet.reset_dimensions()
        sheet.calculate_dimension(force=True)
    return list(sheet.iter_rows(values_only=True))


def read_pdf_sheets(file_stream: BytesIO) -> list[tuple[str, list[tuple]]]:
    file_stream.seek(0)
    rows: list[tuple] = []
    with pdfplumber.open(file_stream) as pdf:
        for page in pdf.pages:
            page_rows = extract_rows_from_pdf_page(page)
            if page_rows:
                rows.extend(page_rows)
            else:
                rows.extend(extract_rows_from_pdf_text(page))
    return [("PDF_Convertido", rows)]


def extract_rows_from_pdf_page(page) -> list[tuple]:
    extracted: list[tuple] = []
    for table in page.extract_tables():
        for row in table:
            if not row:
                continue
            cleaned = tuple(clean_pdf_cell(cell) for cell in row)
            if any(cleaned):
                extracted.append(cleaned)
    return extracted


def extract_rows_from_pdf_text(page) -> list[tuple]:
    text = page.extract_text() or ""
    rows: list[tuple] = []
    for line in text.splitlines():
        parts = [s.strip() for s in re.split(r"\s{2,}", line) if s.strip()]
        if parts:
            rows.append(tuple(parts))
    return rows


def clean_pdf_cell(value: object) -> str:
    return normalize_spaces(value)


# ═══════════════════════════════════════════════════════════════════════════════
# Layout detection
# ═══════════════════════════════════════════════════════════════════════════════

def detect_balancete_layout(rows: list[list]) -> dict | None:
    for index in range(min(len(rows), 10)):
        norm_row = [normalize_text(v) for v in rows[index]]
        columns = {
            "conta": find_header_index(norm_row, {"conta"}),
            "descricao": find_header_index(norm_row, {"descricao", "nome"}),
            "saldo_anterior": find_header_index(norm_row, {"saldo anterior"}),
            "debito": find_header_index(norm_row, {"valor debito", "debito"}),
            "credito": find_header_index(norm_row, {"valor credito", "credito"}),
            "saldo_atual": find_header_index(norm_row, {"saldo atual"}),
        }
        required = {"conta", "descricao", "saldo_anterior", "debito", "credito", "saldo_atual"}
        if all(columns[k] is not None for k in required):
            columns["header_index"] = index
            columns["red"] = None  # optional – not required
            # Optional flag columns present in some TOTVS Balancete exports
            columns["estoque"]      = find_header_index(norm_row, {"estoque", "stk"})
            columns["ativo_imob"]   = find_header_index(norm_row, {"ativo imob", "imobilizado"})
            columns["deprec_acum"]  = find_header_index(norm_row, {"deprec acum", "depreciacao acumulativa", "depreciacao acum"})
            columns["conta_financ"] = find_header_index(norm_row, {"conta financeira"})
            columns["reserva"]      = find_header_index(norm_row, {"reserva de contingencia", "reserva contingencia"})
            return columns
    return None


def detect_diario_layout(rows: list[list]) -> dict | None:
    for index in range(min(len(rows), 15)):
        norm_row = [normalize_text(v) for v in rows[index]]
        columns = {
            "lote": find_header_index(norm_row, {"lote"}),
            "nr_mvto": find_header_index(norm_row, {"nr. mvto", "nr mvto", "nr.mvto"}),
            "conta_debito": find_header_index(norm_row, {"cont. debito", "cont debito", "conta debito"}),
            "conta_credito": find_header_index(norm_row, {"cont. credito", "cont credito", "conta credito"}),
            "historico": find_header_index(norm_row, {"historico"}),
            "debito": find_header_index(norm_row, {"valor debito", "debito"}),
            "credito": find_header_index(norm_row, {"valor credito", "credito"}),
        }
        if all(v is not None for v in columns.values()):
            columns["header_index"] = index
            # Detect optional description columns immediately after account code columns.
            # TOTVS exports often have the account name in the column right after the code.
            detected_set = {v for v in columns.values() if v is not None}
            col_d = columns["conta_debito"]
            col_c = columns["conta_credito"]
            columns["desc_debito"] = (col_d + 1) if (col_d + 1) not in detected_set else None
            columns["desc_credito"] = (col_c + 1) if (col_c + 1) not in detected_set else None
            return columns
    return None


def detect_razao_layout(rows: list[list]) -> dict | None:
    # Search up to 60 rows — TOTVS repeats the account header on each new XLS
    # sheet (when the 65534-row XLS limit is hit), so the header may appear
    # well past row 0 on continuation sheets.
    for index in range(min(len(rows), 60)):
        norm_row = [normalize_text(v) for v in rows[index]]
        if "conta analitica:" in norm_row or "conta analitica" in norm_row:
            return {"account_header_index": index}
    return None


def detect_structured_layout(rows: list[list]) -> dict | None:
    for index in range(min(len(rows), 20)):
        row = rows[index]
        columns: dict[str, int] = {}
        for cell_index, value in enumerate(row):
            norm = normalize_text(value)
            if not norm:
                continue
            for key, aliases in HEADER_ALIASES.items():
                if norm in aliases and key not in columns:
                    columns[key] = cell_index
        if {"date", "description"}.issubset(columns) and columns.keys() & {"debit", "credit", "saldo"}:
            return {"header_index": index, "columns": columns}
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# Balancete parsing
# ═══════════════════════════════════════════════════════════════════════════════

def parse_balancete_rows(rows: list[list], columns: dict) -> list[dict]:
    parsed: list[dict] = []

    for row in rows[columns["header_index"] + 1:]:
        conta = normalize_spaces(get_value(row, columns["conta"]))
        # Skip rows without an account code: Conta Aux. detail rows, footer
        # totals, and blank separator rows are all identified by an empty conta.
        if not conta:
            continue
        conta_norm = normalize_text(conta)
        if conta_norm == "conta":  # header-repetition row
            continue
        if conta_norm.startswith("conta aux"):  # supplier/detail sub-rows
            continue
        if conta_norm.startswith("total") or conta_norm.startswith("subtotal"):
            continue

        # Preserve raw description text (double spaces, etc.) as exported by TOTVS;
        # only strip leading/trailing whitespace and normalise newlines.
        _desc_raw = get_value(row, columns["descricao"])
        descricao = ("" if _desc_raw is None else str(_desc_raw)).replace("\n", " ").replace("\r", " ").strip()
        if normalize_text(descricao) == "descricao":  # header-repetition row
            continue

        # SaldoAnterior: scan BACKWARDS from just before Débito to just after
        # the description column.  This is robust regardless of where
        # detect_balancete_layout placed the "Saldo Anterior" label: the real
        # data column is always the rightmost money cell before Débito.
        # Forward scans fail when a zero-filled filler column sits between the
        # detected label column and the real data column — the first non-None
        # value is 0 and the scan stops too early.
        # Strategy: prefer the first non-zero found while scanning right-to-left;
        # keep the rightmost zero as fallback for accounts whose balance is
        # genuinely zero.
        desc_col = columns["descricao"]
        deb_col  = columns["debito"]
        sa = None
        _sa_zero_fallback: Decimal | None = None
        for _c in range(deb_col - 1, desc_col, -1):
            _v = parse_money_value(get_value(row, _c))
            if _v is None:
                continue
            if _v != Decimal("0"):
                sa = _v
                break
            if _sa_zero_fallback is None:
                _sa_zero_fallback = _v  # rightmost zero = the actual SA column
        if sa is None:
            sa = _sa_zero_fallback
        deb  = parse_money_value(get_value(row, columns["debito"]))
        cred = parse_money_value(get_value(row, columns["credito"]))
        sat  = parse_money_value(get_value(row, columns["saldo_atual"]))

        # Flag columns: read from source when the column is present in the
        # export.  ContaFinanceira has one universal rule: it applies only to
        # leaf sub-accounts (prefix ends with a dot before the last segment),
        # never to aggregator/parent groups.  Estoque and ReservaDeContingência
        # vary per hospital unit and must come from the source — no hardcoded
        # prefix rules are used for those two flags.
        def _flag(col_key: str) -> str:
            col = columns.get(col_key)
            if col is not None:
                v = normalize_text(get_value(row, col))
                return "S" if v == "s" else "N"
            # ContaFinanceira rule: S only for leaf accounts under the two
            # financial sub-groups (trailing dot ensures parent groups are
            # excluded: "1.1.1.02.02" does not start with "1.1.1.02.02.").
            if col_key == "conta_financ":
                return "S" if (
                    conta.startswith("1.1.1.02.02.") or conta.startswith("1.1.1.02.04.")
                ) else "N"
            # ReservaDeContingência: when no source column exists, detect by
            # account name keywords.  Reserve/contingency accounts within the
            # financial sub-groups consistently contain "fundo", "reserva", or
            # "contingencia" per IMED naming convention (unit-independent).
            if col_key == "reserva":
                if conta.startswith("1.1.1.02.02.") or conta.startswith("1.1.1.02.04."):
                    nome_lower = normalize_text(descricao)
                    return "S" if any(kw in nome_lower for kw in ("fundo", "reserva", "contingencia")) else "N"
                return "N"
            return "N"

        parsed.append({
            "REG": "0300",
            "ContaContabil": conta,
            "NomeConta": descricao or "Sem descricao",
            "SaldoAnterior": sa,
            "Debito": deb,
            "Credito": cred,
            "SaldoAtual": sat,
            "Estoque": _flag("estoque"),
            "AtivoImob": _flag("ativo_imob"),
            "DepreciacaoAcumulativa": _flag("deprec_acum"),
            "ContaFinanceira": _flag("conta_financ"),
            "ReservaDeContingência": _flag("reserva"),
            "Centro de Custos": None,
        })

    return parsed


def _extract_balancete_money(
    row: list,
) -> tuple[Decimal | None, Decimal | None, Decimal | None, Decimal | None]:
    money_values = [parse_money_value(v) for v in row if parse_money_value(v) is not None]
    if len(money_values) >= 4:
        return money_values[-4], money_values[-3], money_values[-2], money_values[-1]
    if len(money_values) == 3:
        return None, money_values[0], money_values[1], money_values[2]
    if len(money_values) == 2:
        return None, money_values[0], money_values[1], None
    if len(money_values) == 1:
        return None, None, None, money_values[0]
    return None, None, None, None


# ═══════════════════════════════════════════════════════════════════════════════
# Diário parsing
# ═══════════════════════════════════════════════════════════════════════════════

def parse_diario_rows(rows: list[list], columns: dict) -> list[dict]:
    """
    Process ALL non-empty rows (including date headers and repeated column headers).
    Each XLS sub-row (debit side OR credit side) becomes one HCN output row.
    Combined rows (both sides in one row) are split into two output rows.

    TOTVS sometimes puts the debit account on one row and the credit account on the
    next row, omitting the Histórico and the account name on the secondary row.
    We carry those forward from the most recent non-empty values.
    """
    parsed: list[dict] = []
    current_date = ""
    last_historico = ""          # carry-forward: TOTVS omits historico on credit-only rows
    last_debito_name = ""        # carry-forward: debit account name for complement rows
    last_credito_name = ""       # carry-forward: credit account name for complement rows

    col_lote = columns["lote"]
    col_nr_mvto = columns.get("nr_mvto")
    col_debito_conta = columns["conta_debito"]
    col_credito_conta = columns["conta_credito"]
    col_historico = columns["historico"]
    col_debito_val = columns["debito"]
    col_credito_val = columns["credito"]
    # Optional: column immediately after each account code column may hold the account name
    col_debito_desc = columns.get("desc_debito")
    col_credito_desc = columns.get("desc_credito")

    for row in rows:
        texts = [normalize_spaces(v) for v in row]

        # ── Date header row: "01 de Março de 2026" in one of the cells ──
        for cell_val in texts:
            pt_date = parse_pt_date_header(cell_val)
            if pt_date:
                current_date = pt_date
                break
        else:
            # Not a date header – process as data row
            lote_text = _safe_col(texts, col_lote)

            # Skip column-header repetition rows
            if normalize_text(lote_text) == "lote":
                continue

            # Skip continuation/footer rows that have no lote value
            if not lote_text:
                continue

            conta_debito = _safe_col(texts, col_debito_conta)
            conta_credito = _safe_col(texts, col_credito_conta)

            # Histórico: use carry-forward when the current row's column is empty.
            # TOTVS only writes the historico on the main (first) row of each movement.
            raw_historico = _safe_col(texts, col_historico)
            if raw_historico:
                last_historico = raw_historico
            historico = raw_historico or last_historico or "Sem historico"

            debito_val = parse_money_value(_safe_col(texts, col_debito_val))
            credito_val = parse_money_value(_safe_col(texts, col_credito_val))

            # Nr.Mvto — same value for both the debit and credit rows of one lançamento.
            mov_val = None
            if col_nr_mvto is not None:
                raw_mov = _safe_col(texts, col_nr_mvto)
                if raw_mov:
                    try:
                        mov_val = int(float(raw_mov))
                    except (ValueError, TypeError):
                        mov_val = raw_mov

            # Continuation rows: account code must start with a digit
            if conta_debito and not re.match(r'^\d', conta_debito.strip()):
                conta_debito = ""
            if conta_credito and not re.match(r'^\d', conta_credito.strip()):
                conta_credito = ""

            if not conta_debito and not conta_credito:
                continue

            # Normalize double spaces in Diário text fields (fix for 42-line
            # divergence; Razão intentionally preserves them via _extract_razao_hist_raw).
            historico_out = ' '.join(historico.split())

            # Build both sides of the movement, then apply HCN ordering:
            # group 1.x/2.x (ativo/passivo) always precedes group 3.x
            # (despesa/receita) within the same Mov pair.
            pending: list[dict] = []

            if conta_credito:
                code, name = split_account_field(conta_credito)
                if not name and col_credito_desc is not None:
                    adj = _safe_col(texts, col_credito_desc)
                    if adj and parse_money_value(adj) is None:
                        name = adj
                if not name:
                    name = last_credito_name
                elif name:
                    last_credito_name = name
                pending.append({
                    "REG": 1600,
                    "DATA": current_date,
                    "CLASSIFICAÇÃO": code,
                    "DESCRIÇÃO": ' '.join(name.split()),
                    "HISTÓRICO": historico_out,
                    "DÉBITO": None,
                    "CRÉDITO": decimal_to_float(credito_val),
                    "Centro de Custos": None,
                    "Mov": mov_val,
                })

            if conta_debito:
                code, name = split_account_field(conta_debito)
                if not name and col_debito_desc is not None:
                    adj = _safe_col(texts, col_debito_desc)
                    if adj and parse_money_value(adj) is None:
                        name = adj
                if not name:
                    name = last_debito_name
                elif name:
                    last_debito_name = name
                pending.append({
                    "REG": 1600,
                    "DATA": current_date,
                    "CLASSIFICAÇÃO": code,
                    "DESCRIÇÃO": ' '.join(name.split()),
                    "HISTÓRICO": historico_out,
                    "DÉBITO": decimal_to_float(debito_val),
                    "CRÉDITO": 0.0,
                    "Centro de Custos": None,
                    "Mov": mov_val,
                })

            if len(pending) == 2:
                # Both sides present: group 1/2 before group 3.
                # pending[0] = crédito row, pending[1] = débito row (built above).
                grp_cre = str(pending[0]["CLASSIFICAÇÃO"]).split('.')[0]
                grp_deb = str(pending[1]["CLASSIFICAÇÃO"]).split('.')[0]
                if grp_deb in ('1', '2') and grp_cre == '3':
                    # Exception: débito is 1/2, crédito is 3 → débito goes first
                    parsed.append(pending[1])
                    parsed.append(pending[0])
                else:
                    # Default: crédito first (covers 1/2-cre + 3-deb and same-group pairs)
                    parsed.append(pending[0])
                    parsed.append(pending[1])
            else:
                parsed.extend(pending)

    # Post-process: fix ordering for split movements (each side on its own
    # source row).  The combined-row branch above already handles combined
    # rows; here we catch consecutive same-Mov pairs where group 3 ended up
    # before group 1/2 because TOTVS writes the débito row first.
    result: list[dict] = []
    i = 0
    while i < len(parsed):
        if (
            i + 1 < len(parsed)
            and parsed[i]["Mov"] == parsed[i + 1]["Mov"]
            and str(parsed[i]["CLASSIFICAÇÃO"]).split(".")[0] == "3"
            and str(parsed[i + 1]["CLASSIFICAÇÃO"]).split(".")[0] in ("1", "2")
        ):
            result.append(parsed[i + 1])  # group 1/2 (ativo/passivo) first
            result.append(parsed[i])      # group 3  (despesa/receita) second
            i += 2
        else:
            result.append(parsed[i])
            i += 1
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Razão parsing helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_razao_hist(texts: list[str]) -> str:
    """Return the normalised historico text (used only for detection checks)."""
    for ci in (19, 20, 21):
        v = _safe_col(texts, ci).strip()
        if v:
            return v
    return ""


def _extract_razao_hist_raw(row: list) -> str:
    """Return the historico text preserving original spacing (double spaces, leading space)."""
    for ci in (19, 20, 21):
        if ci < len(row) and row[ci] is not None:
            v = str(row[ci])
            if v.strip():
                return v
    return ""


def _find_saldo_anterior(texts: list[str]) -> Decimal | None:
    """Locate the Saldo Anterior value in an account-header row.

    The label "Saldo Anterior:" and its value appear at different column
    positions across TOTVS export versions (e.g. cols 20/22 or 22/24).
    We search for the label and read the first parseable money value that
    follows it within a small window.
    """
    for ci, tv in enumerate(texts):
        if normalize_text(tv) in {"saldo anterior:", "saldo anterior"}:
            for cj in range(ci + 1, min(len(texts), ci + 5)):
                v = parse_money_value(texts[cj])
                if v is not None:
                    return v
            break
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# Razão parsing
# ═══════════════════════════════════════════════════════════════════════════════

def parse_razao_rows(rows: list[list], layout: dict) -> list[dict]:
    parsed: list[dict] = []
    current_account_name = ""
    current_account_code = ""
    current_saldo_anterior: Decimal | None = None
    running_saldo: Decimal | None = None   # Bug 2: accumulated per account
    current_date = ""
    pending_record: dict | None = None
    expecting_contra = False
    pending_historico = ""  # historico lives on a standalone row before the transaction row

    start_idx = layout["account_header_index"]
    # When a sheet begins with continuation rows (before its first account
    # header), pre-populate the account context from that header so those
    # rows receive the correct account code and name.
    if start_idx > 0:
        pre = rows[start_idx]
        pre_texts = [normalize_spaces(v) for v in pre]
        if _safe_col(pre_texts, 1) in {"Conta Analitica:", "Conta Analítica:"}:
            current_account_code, current_account_name = parse_razao_account_header(
                _safe_col(pre_texts, 5)
            )
            current_saldo_anterior = _find_saldo_anterior(pre_texts)
            running_saldo = current_saldo_anterior
        # Pre-populate the current date from the first date found in this sheet
        # so that continuation rows (before the account header) get a valid date.
        for _pre_row in rows:
            _pre_date = parse_date_value(_safe_col([normalize_spaces(v) for v in _pre_row], 1))
            if _pre_date:
                current_date = _pre_date
                break

    for row_idx, row in enumerate(rows):
        texts = [normalize_spaces(v) for v in row]
        non_empty_vals = [v for v in texts if v]

        if not non_empty_vals:
            continue

        col1  = _safe_col(texts, 1)
        col3  = _safe_col(texts, 3)
        col5  = _safe_col(texts, 5)
        col6  = _safe_col(texts, 6)
        col8  = _safe_col(texts, 8)
        col10 = _safe_col(texts, 10)
        col13 = _safe_col(texts, 13)
        col_hist     = _extract_razao_hist(texts)       # normalised — for detection checks
        col_hist_raw = _extract_razao_hist_raw(row)     # raw — stored value preserves spaces

        # ── (1) Account header row (always checked first) ──────────────────
        if col1 in {"Conta Analitica:", "Conta Analítica:"}:
            expecting_contra = False
            new_code, new_name = parse_razao_account_header(col5)
            # When TOTVS breaks a page it reprints the same account header.
            # If the account code changes we start a fresh context; if it is the
            # same account (repeated page header) we keep pending_record and
            # pending_historico so the Contrapartida/historico on the next page
            # can still be linked to the transaction on the previous page.
            if new_code != current_account_code:
                pending_record = None
                pending_historico = ""
                # Bug 3: find saldo anterior dynamically — label position varies
                # across TOTVS export versions (col 20 or col 22 in different files).
                current_saldo_anterior = _find_saldo_anterior(texts)
                running_saldo = current_saldo_anterior  # reset per-account accumulator
            current_account_code = new_code
            current_account_name = new_name
            continue

        # ── (2) Account code row after Contrapartida: ──────────────────────
        if expecting_contra:
            if parse_date_value(col1) is not None or col3 in {"Contrapartida:", "Contrapartida"}:
                expecting_contra = False
                # fall through to date/transaction handling below
            else:
                # Only store the FIRST contra partida; subsequent ones are ignored.
                if col5 and pending_record is not None and not pending_record["Contra Partida"]:
                    pending_record["Contra Partida"] = parse_razao_contra_code(col5)
                # B1: when there is no blank line between the Contrapartida block and
                # the next historico row, the historico row is consumed here.  Capture it
                # so the following transaction can still use it.
                if col_hist and not col5.strip():
                    pending_historico = col_hist_raw
                expecting_contra = False
                continue

        # ── (3) Skip "Data" column header ─────────────────────────────────
        if col1 == "Data" or (normalize_text(col1) == "data" and parse_date_value(col1) is None):
            continue

        # ── (4) Contrapartida label row ────────────────────────────────────
        if col3 in {"Contrapartida:", "Contrapartida"}:
            # Some TOTVS formats put the contra account code on the SAME row
            # as the label (col5); capture it here before continuing.
            if col5 and pending_record is not None and not pending_record["Contra Partida"]:
                pending_record["Contra Partida"] = parse_razao_contra_code(col5)
            # Retroactively improve the historico of the current record when the
            # Contrapartida row carries a longer col19 text.  This fixes B1 blocks
            # where the transaction row received a truncated historico (no trailing
            # space, so _completar_historico was never called) but the CP row has
            # the complete version.  Also fixes CTR double-space cases.
            if pending_record is not None and col_hist_raw.strip():
                current_hist = pending_record.get("HISTÓRICO") or ""
                if current_hist in ("Sem historico", "") or \
                        len(col_hist_raw.strip()) > len(current_hist.strip()):
                    pending_record["HISTÓRICO"] = col_hist_raw.strip()
            # A real Contrapartida row always has monetary values in col6/col10.
            # Rows with only text (no amounts) are page-break continuation
            # fragments of the previous historico and must NOT trigger the flag.
            if parse_money_value(col6) is not None or parse_money_value(col10) is not None:
                expecting_contra = True
            continue

        debit  = parse_money_value(col8)
        credit = parse_money_value(col13)

        # ── (4b) Standalone historico row (text only, no date, no amounts) ─
        # TOTVS places the historico text in col 19, 20 or 21 on a dedicated row
        # BEFORE the transaction row.  Store the raw value to preserve spacing.
        if col_hist and debit is None and credit is None and not parse_date_value(col1):
            pending_historico = col_hist_raw
            continue

        # ── (5) Transaction row ────────────────────────────────────────────
        row_date = parse_date_value(col1)
        if row_date:
            current_date = row_date

        if debit is not None or credit is not None:
            historico = pending_historico or col_hist_raw
            pending_historico = ""
            # B2: if historico ends with a trailing space it was truncated in the source.
            # Look ahead for a Contrapartida col19 or an isolated 2nd-part row.
            if historico.endswith(' '):
                historico = _completar_historico(rows, row_idx, historico)
            d = debit  if debit  is not None else Decimal(0)
            c = credit if credit is not None else Decimal(0)
            saldo_base = running_saldo if running_saldo is not None else Decimal(0)
            running_saldo = _calcular_saldo(saldo_base, d, c, current_account_code, current_account_name)
            pending_record = {
                "REG": 1700,
                "NOME CONTA": current_account_name or "Sem conta",
                "CONTA CONTÁBIL": current_account_code or "",
                "SALDO ANTERIOR": decimal_to_float(current_saldo_anterior),
                "DATA": current_date,
                "HISTÓRICO": historico if historico.strip() else "Sem historico",
                "DÉBITO": decimal_to_float(debit),
                "CRÉDITO": decimal_to_float(credit),
                "Saldo": decimal_to_float(running_saldo),
                "Contra Partida": "",
                "Centro de Custos": None,
            }
            parsed.append(pending_record)

    return parsed


# ═══════════════════════════════════════════════════════════════════════════════
# SOULMV Razão parsing (e.g. hospital HCN — layout "R_RAZAO_RET_SC")
# ═══════════════════════════════════════════════════════════════════════════════
# This export is a printed report captured cell-by-cell: every account renders
# as a repeating block of merged-cell rows, and a page break re-wraps the
# layout so the debit/credit/saldo values do not sit in a fixed column from
# block to block. Rows are classified by their textual markers ("Conta
# Analítica:", "Contrapartida:", "<<== Saldo da Conta", the "Data" column
# header) instead of by column letter, and money values inside a classified
# row are read in left-to-right order (Débito, Crédito[, Saldo]).

_SOULMV_MONEY_CORE_RE = re.compile(r"^\d{1,3}(?:\.\d{3})*,\d{2}$")
_SOULMV_CNPJ_RE = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}")
_SOULMV_CONTA_LABELS = {"Conta Analítica:", "Conta Analitica:"}
_SOULMV_SALDO_ANT_LABELS = {"Saldo Anterior:", "Saldo Anterior"}
_SOULMV_CONTRAPARTIDA_LABELS = {"Contrapartida:", "Contrapartida"}
_SOULMV_HEADER_VALUE_LABELS = {"Valor Débito", "Valor Crédito", "Saldo", "Histórico Padrão"}


def detect_soulmv_razao_layout(rows: list[list], sheet_title: str = "") -> dict | None:
    """Detect the SOULMV 'Razão Contábil' export (e.g. hospital HCN).

    SOULMV prints its own name in the report header ("SOULMV - Sistema de
    Contabilidade"). That text is a reliable, hospital-independent signature
    that never appears in the TOTVS-based Razão exports (IMED Formosa /
    Policlínica Posse) handled by detect_razao_layout/parse_razao_rows, whose
    fixed column offsets would otherwise misparse this layout.

    Some SOULMV exports are printed from a page range that excludes the
    report's cover header (no "SOULMV" text anywhere in the file — e.g.
    LIVRO_RAZAO_*_IMED_HET.xls), so that text can be entirely absent even
    though the block-print layout is identical. SOULMV's internal worksheet
    name ("R_RAZAO_RET_SC", possibly suffixed on continuation sheets) is a
    second, independent signature from the accounting system itself rather
    than the print range, and is used as a fallback when the header text is
    missing.
    """
    has_soulmv = normalize_text(sheet_title).startswith("r_razao_ret_sc")
    account_header_index = None
    for index in range(min(len(rows), 60)):
        norm_row = [normalize_text(v) for v in rows[index]]
        if not has_soulmv and any(v and "soulmv" in v for v in norm_row):
            has_soulmv = True
        if account_header_index is None and (
            "conta analitica:" in norm_row or "conta analitica" in norm_row
        ):
            account_header_index = index
        if has_soulmv and account_header_index is not None:
            return {"account_header_index": account_header_index}
    return None


def _soulmv_parse_money(text: str) -> Decimal | None:
    # SOULMV writes every value as pt-BR text with a 2-decimal comma
    # ("2.000.000,00"), occasionally with a TRAILING minus for negative
    # balances ("402.666,64-") instead of a leading one. A dedicated, strict
    # parser is used here (instead of the shared parse_money_value, whose
    # bare-integer fallback exists for native numeric cells elsewhere in the
    # app) so stray digits in this text-only export — e.g. the "2" in a
    # "Página: 2" boilerplate row — are never mistaken for a money value.
    if not text:
        return None
    t = text.strip()
    if not t:
        return None
    neg = t.startswith("-") or t.endswith("-")
    core = t.strip("-")
    if not _SOULMV_MONEY_CORE_RE.match(core):
        return None
    try:
        value = Decimal(core.replace(".", "").replace(",", "."))
    except InvalidOperation:
        return None
    return -value if neg else value


def _soulmv_is_boilerplate_line(text: str) -> bool:
    norm = normalize_text(text)
    if norm.startswith("periodo:"):
        return True
    return bool(_SOULMV_CNPJ_RE.search(text))


def parse_soulmv_razao_rows(rows: list[list], layout: dict) -> list[dict]:
    parsed: list[dict] = []
    current_account_code = ""
    current_account_name = ""
    current_saldo_anterior: Decimal | None = None
    current_date = ""
    pending_historico = ""
    pending_record: dict | None = None

    total_rows = len(rows)
    row_index = 0
    while row_index < total_rows:
        row = rows[row_index]
        texts = [normalize_spaces(v) for v in row]
        non_empty = [t for t in texts if t]

        if not non_empty:
            row_index += 1
            continue

        # ── (1) Account header row ──────────────────────────────────────
        if any(t in _SOULMV_CONTA_LABELS for t in texts):
            candidates = [
                t for t in non_empty
                if t not in _SOULMV_CONTA_LABELS
                and t not in _SOULMV_SALDO_ANT_LABELS
                and _soulmv_parse_money(t) is None
            ]
            new_code, new_name = ("", "")
            if len(candidates) == 1:
                new_code, new_name = parse_razao_account_header(candidates[0])
            if new_code and new_code != current_account_code:
                current_saldo_anterior = _find_saldo_anterior(texts)
                pending_record = None
                pending_historico = ""
            if new_code:
                current_account_code = new_code
                current_account_name = new_name
            row_index += 1
            continue

        # ── (2) End-of-account checksum row ("<<== Saldo da Conta: ...") ─
        if any("<<==" in t for t in texts):
            row_index += 1
            continue

        # ── (3) Column header row (repeats on every printed page) ────────
        if "Data" in texts and any(t in _SOULMV_HEADER_VALUE_LABELS for t in texts):
            row_index += 1
            continue

        # ── (4) Contrapartida row ─────────────────────────────────────────
        if any(t in _SOULMV_CONTRAPARTIDA_LABELS for t in texts):
            values = [v for v in (_soulmv_parse_money(t) for t in texts) if v is not None]
            # A real Contrapartida row always carries its own Débito/Crédito.
            # Rows with the label but no amounts are page-break historico
            # fragments and must not be treated as a Contrapartida block.
            if len(values) >= 2 and pending_record is not None:
                next_index = row_index + 1
                while next_index < total_rows and all(
                    normalize_spaces(v) == "" for v in rows[next_index]
                ):
                    next_index += 1
                if next_index < total_rows:
                    next_texts = [normalize_spaces(v) for v in rows[next_index]]
                    next_non_empty = [t for t in next_texts if t]
                    if len(next_non_empty) == 1 and CONTRA_RE.match(next_non_empty[0]):
                        code = parse_razao_contra_code(next_non_empty[0])
                        existing = pending_record.get("Contra Partida") or ""
                        pending_record["Contra Partida"] = (
                            f"{existing}; {code}" if existing else code
                        )
            row_index += 1
            continue

        # ── (5) Transaction / value row ───────────────────────────────────
        values = [v for v in (_soulmv_parse_money(t) for t in texts) if v is not None]
        if values:
            for t in texts:
                if DATE_RE.match(t):
                    current_date = t
                    break
            debito = values[0] if len(values) > 0 else None
            credito = values[1] if len(values) > 1 else None
            saldo = values[2] if len(values) > 2 else None
            historico = pending_historico or "Sem historico"
            pending_historico = ""
            pending_record = {
                "REG": 1700,
                "NOME CONTA": current_account_name or "Sem conta",
                "CONTA CONTÁBIL": current_account_code or "",
                "SALDO ANTERIOR": decimal_to_float(current_saldo_anterior),
                "DATA": current_date,
                "HISTÓRICO": historico,
                "DÉBITO": decimal_to_float(debito),
                "CRÉDITO": decimal_to_float(credito),
                "Saldo": decimal_to_float(saldo),
                "Contra Partida": "",
                "Centro de Custos": None,
            }
            parsed.append(pending_record)
            row_index += 1
            continue

        # ── (6) Standalone text row: historico, contra-account code (the
        #     code row is consumed via the lookahead in step 4 above; when
        #     the main loop reaches it independently, do nothing), or
        #     report boilerplate (period line, hospital name + CNPJ) ────────
        if len(non_empty) == 1:
            text = non_empty[0]
            if not _soulmv_is_boilerplate_line(text) and not CONTRA_RE.match(text):
                pending_historico = text
            row_index += 1
            continue

        row_index += 1

    return parsed


# ═══════════════════════════════════════════════════════════════════════════════
# Generic / structured fallback parsers (unchanged)
# ═══════════════════════════════════════════════════════════════════════════════

def parse_structured_rows(
    rows: list[list], header_index: int, columns: dict
) -> list[dict]:
    records: list[dict] = []
    current: dict | None = None
    ignored = set(columns.values())

    for row in rows[header_index + 1:]:
        parsed_date = parse_date_value(get_value(row, columns.get("date")))
        description = normalize_spaces(get_value(row, columns.get("description")))
        debit = parse_money_value(get_value(row, columns.get("debit")))
        credit = parse_money_value(get_value(row, columns.get("credit")))
        saldo = parse_money_value(get_value(row, columns.get("saldo")))

        if parsed_date:
            current = {
                "Data": parsed_date, "Descricao": description,
                "Debito": debit, "Credito": credit, "Saldo": saldo,
            }
            append_extra_description(current, row, ignored)
            records.append(current)
            continue

        if current is None:
            continue

        continued = description or collect_text_fragments(row, ignored)
        if continued:
            current["Descricao"] = join_description(current["Descricao"], continued)

        if current["Debito"] is None and debit is not None:
            current["Debito"] = debit
        if current["Credito"] is None and credit is not None:
            current["Credito"] = credit
        if current["Saldo"] is None and saldo is not None:
            current["Saldo"] = saldo

    return finalize_records(records)


def parse_generic_rows(rows: list[list]) -> list[dict]:
    records: list[dict] = []
    current: dict | None = None

    for row in rows:
        parsed_date = first_date_in_row(row)
        money_cells = extract_money_cells(row)
        text_parts = extract_text_parts(row)

        if parsed_date:
            debit, credit, saldo = distribute_amounts(money_cells)
            current = {
                "Data": parsed_date, "Descricao": " ".join(text_parts),
                "Debito": debit, "Credito": credit, "Saldo": saldo,
            }
            records.append(current)
            continue

        if current is None:
            continue

        continuation = " ".join(text_parts)
        if continuation:
            current["Descricao"] = join_description(current["Descricao"], continuation)

        debit, credit, saldo = distribute_amounts(money_cells)
        if current["Debito"] is None and debit is not None:
            current["Debito"] = debit
        if current["Credito"] is None and credit is not None:
            current["Credito"] = credit
        if current["Saldo"] is None and saldo is not None:
            current["Saldo"] = saldo

    return finalize_records(records)


def finalize_records(records: list[dict]) -> list[dict]:
    cleaned = []
    for r in records:
        if not r["Data"]:
            continue
        r["Descricao"] = normalize_spaces(r["Descricao"]) or "Sem descricao"
        cleaned.append(r)
    return cleaned


# ═══════════════════════════════════════════════════════════════════════════════
# Output writing & styling
# ═══════════════════════════════════════════════════════════════════════════════

def write_records(sheet, parsed_sheet: dict) -> None:
    headers = parsed_sheet["headers"]
    sheet.append(headers)
    for record in parsed_sheet["rows"]:
        row = []
        for header in headers:
            value = record.get(header)
            row.append(decimal_to_float(value) if isinstance(value, Decimal) else value)
        sheet.append(row)


def style_output_sheet(sheet, parsed_sheet: dict, sheet_index: int) -> None:
    headers = parsed_sheet["headers"]
    last_row = sheet.max_row
    last_col = sheet.max_column
    money_columns = parsed_sheet["money_columns"]
    desc_col = parsed_sheet["description_column"]

    is_hcn = headers in (BALANCETE_HCN_HEADERS, DIARIO_HCN_HEADERS, RAZAO_HCN_HEADERS)

    if is_hcn:
        # HCN: céluals planas, sem tabela, sem freeze — replicar estrutura do modelo de referência
        if headers != DIARIO_HCN_HEADERS:
            sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in range(2, last_row + 1):
            for col in range(1, last_col + 1):
                cell = sheet.cell(row=row, column=col)
                if col in money_columns:
                    if headers == RAZAO_HCN_HEADERS:
                        # SALDO ANTERIOR (col 4): R$ + zero como traço
                        # DÉBITO/CRÉDITO/Saldo (cols 7/8/9): apenas zero como traço
                        cell.number_format = (
                            ACCOUNTING_FORMAT_BRL if col == 4 else ACCOUNTING_FORMAT_DASH
                        )
                    else:
                        cell.number_format = ACCOUNTING_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col == desc_col:
                    cell.alignment = Alignment(wrap_text=True, vertical="center")

    else:
        # Genérico: estilo decorado com tabela, freeze e preenchimento alternado
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        date_columns = parsed_sheet["date_columns"]

        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        for row in range(2, last_row + 1):
            for col in range(1, last_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=(col == desc_col))

                if row % 2 == 0:
                    cell.fill = ALT_ROW_FILL

                if col in date_columns:
                    cell.number_format = "dd/mm/yyyy"
                elif col in money_columns:
                    cell.number_format = ACCOUNTING_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        highlight_total_rows(sheet, last_row, last_col, desc_col)
        create_table(sheet, last_row, last_col, sheet_index)

    adjust_column_widths(sheet, parsed_sheet)


def highlight_total_rows(sheet, last_row: int, last_col: int, desc_col: int) -> None:
    for row in range(2, last_row + 1):
        desc = sheet.cell(row=row, column=desc_col).value
        if isinstance(desc, str) and any(
            kw in desc.lower() for kw in ("total", "subtotal", "saldo anterior", "resumo")
        ):
            for col in range(1, last_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = Font(bold=True, color="1F1F1F")
                cell.fill = PatternFill("solid", fgColor="D9EAD3")
                cell.border = TOTAL_BORDER


def adjust_column_widths(sheet, parsed_sheet: dict) -> None:
    headers = parsed_sheet["headers"]
    if headers == BALANCETE_HCN_HEADERS:
        widths = {1: 8, 2: 22, 3: 54, 4: 18, 5: 18, 6: 18, 7: 18, 8: 10, 9: 12, 10: 22, 11: 16, 12: 22, 13: 16}
    elif headers == DIARIO_HCN_HEADERS:
        widths = {1: 9, 2: 11, 3: 22, 4: 57, 5: 96, 6: 18, 7: 18, 8: 22, 9: 15}
    elif headers == RAZAO_HCN_HEADERS:
        # REG NOME_CONTA CONTA SALDO_ANT DATA HISTÓRICO DÉBITO CRÉDITO Saldo ContraP CentroCusto
        widths = {1: 8, 2: 44, 3: 22, 4: 18, 5: 16, 6: 56, 7: 18, 8: 18, 9: 18, 10: 22, 11: 18}
    else:
        widths = {1: 14, 2: 60, 3: 16, 4: 16, 5: 16}
    for col_idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def create_table(sheet, last_row: int, last_col: int, sheet_index: int) -> None:
    if last_row < 2:
        return
    table_range = f"A1:{get_column_letter(last_col)}{last_row}"
    safe_title = re.sub(r"[^A-Za-z0-9_]", "", sheet.title)[:18] or "Planilha"
    table = Table(
        displayName=f"Tabela_{sheet_index + 1}_{safe_title}",
        ref=table_range,
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    sheet.add_table(table)


def build_sheet_title(headers: list, index: int) -> str:
    if headers == BALANCETE_HCN_HEADERS:
        return "0300"
    if headers == DIARIO_HCN_HEADERS:
        return "1600"
    if headers == RAZAO_HCN_HEADERS:
        return "1700"
    return f"Planilha{index + 1}"


# ═══════════════════════════════════════════════════════════════════════════════
# Helper / utility functions
# ═══════════════════════════════════════════════════════════════════════════════

def _completar_historico(rows: list[list], start_idx: int, hist: str) -> str:
    """Tenta completar um histórico truncado (termina com espaço) via Contrapartidas.

    Varre as linhas de Contrapartida dentro do mesmo bloco.
    Condições de parada:
      1. Nova DATA em col1 → início de novo lançamento.
      2. "Conta Analítica:" em col1 → cabeçalho de nova conta.
      3. Após ao menos uma CP encontrada (cp_found), qualquer linha com col19
         preenchida e sem marcador de Contrapartida → é o histórico antecipado
         do próximo lançamento; não deve ser capturado.
    A condição 3 impede que a busca ultrapasse o código da conta (col5) do bloco
    atual e chegue à linha de histórico do bloco seguinte, que fica 1-2 linhas
    após o último código de contra-partida.
    """
    hist_base = hist.rstrip()
    best = hist_base
    cp_found = False

    for offset in range(1, 13):
        idx = start_idx + offset
        if idx >= len(rows):
            break
        ahead_texts = [normalize_spaces(v) for v in rows[idx]]
        ahead_col1 = _safe_col(ahead_texts, 1)
        ahead_col3 = _safe_col(ahead_texts, 3)

        # Parar ao cruzar fronteira do bloco (condições 1 e 2)
        if parse_date_value(ahead_col1) is not None:
            break
        if ahead_col1 in {"Conta Analitica:", "Conta Analítica:"}:
            break

        if ahead_col3 in {"Contrapartida:", "Contrapartida"}:
            cp_found = True
            cp = _extract_razao_hist_raw(rows[idx]).strip()
            if cp and len(cp) > len(best):
                best = cp
            # Continuar — pode haver mais CPs no bloco (ex: 2 CPs para o mesmo lançamento)
        elif cp_found and _extract_razao_hist_raw(rows[idx]).strip():
            # Condição 3: linha com col19 preenchida após o bloco de CPs
            # = histórico do próximo lançamento → parar imediatamente.
            break

    return best if best else hist_base


def _calcular_saldo(saldo_ant: Decimal, d: Decimal, c: Decimal,
                    conta: str, nome: str = '') -> Decimal:
    """Acumula saldo respeitando a natureza contábil da conta.

    Passivo (grupo 2) e Receitas (3.2.1.01.x): crédito aumenta → ant - D + C.
    Ativo (grupo 1) e Despesas: débito aumenta → ant + D - C.
    Contas retificadoras (nome começa com "(-)") invertem a lógica do grupo.
    """
    grupo = conta.strip().split('.')[0] if conta.strip() else ''
    is_retificadora = nome.strip().upper().startswith('(-)')
    natureza_cd = (grupo == '2' or conta.strip().startswith('3.2.1.01'))
    if natureza_cd ^ is_retificadora:   # XOR: retificadora inverte a natureza
        return saldo_ant - d + c
    return saldo_ant + d - c


def parse_pt_date_header(text: str) -> str | None:
    """Parse 'DD de Mês de AAAA' (from TOTVS date headers) → 'DD/MM'."""
    if not text:
        return None
    norm = normalize_text(text)  # removes accents, lowercases
    m = re.match(r"^(\d{1,2})\s+de\s+(\w+)\s+de\s+\d{4}$", norm)
    if m:
        day = m.group(1).zfill(2)
        month = PT_MONTHS.get(m.group(2))
        if month:
            return f"{day}/{month}"
    return None


def split_account_field(text: str) -> tuple[str, str]:
    """Split 'CODE - NAME' into (code, name). Handles trailing dash with no name."""
    stripped = text.strip()
    # Full format: "CODE - NAME"
    m = re.match(r"^([\d.]+)\s*-\s*(.+)$", stripped)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    # Code only, with or without trailing dash: "CODE -" or "CODE"
    m2 = re.match(r"^([\d.]+)\s*-?\s*$", stripped)
    if m2:
        return m2.group(1).strip(), ""
    return stripped, stripped


def parse_razao_account_header(text: str) -> tuple[str, str]:
    """Parse 'SEQNO CODE - NAME' → (code, name). E.g. '92576 1.1.1.02.02.010 - BCO CEF...'"""
    if not text:
        return "", ""
    m = ACCOUNT_RE.match(text.strip())
    if m:
        return m.group(2).strip(), m.group(3).strip()
    # Fallback: try CODE - NAME without sequence
    m2 = re.match(r"^([\d.]+)\s*-\s*(.+)$", text.strip())
    if m2:
        return m2.group(1).strip(), m2.group(2).strip()
    return "", text.strip()


def parse_razao_contra_code(text: str) -> str:
    """Parse 'SEQNO - CODE' → 'CODE'. E.g. '10028 - 2.1.1.02.01.001' → '2.1.1.02.01.001'"""
    if not text:
        return ""
    m = CONTRA_RE.match(text.strip())
    if m:
        return m.group(2).strip()
    # Maybe already just a code
    return text.strip()


def _safe_col(texts: list[str], index: int) -> str:
    return texts[index] if index < len(texts) else ""


def row_is_empty(row: tuple | list) -> bool:
    return all(normalize_spaces(v) == "" for v in row)


def normalize_text(value: object) -> str:
    text = normalize_spaces(value)
    if not text:
        return ""
    text = (
        unicodedata.normalize("NFKD", text)
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )
    return re.sub(r"\s+", " ", text).strip()


def normalize_spaces(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def parse_date_value(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    text = normalize_spaces(value)
    if not text:
        return None
    text = text.split(" ")[0]

    if DATE_RE.match(text):
        try:
            return datetime.strptime(text, "%d/%m/%Y").strftime("%d/%m/%Y")
        except ValueError:
            return None

    for fmt in ("%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue

    return None


_US_NUMBER_RE = re.compile(r"^-?\d+(?:\.\d+)?$")


def parse_money_value(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = normalize_spaces(value)
    if not text:
        return None

    # XLS/XLSX numeric cells converted to string via normalize_spaces() arrive in
    # US dot-decimal format ("1234.56", "1200.0").  Handle these before the BR regex.
    if _US_NUMBER_RE.match(text):
        try:
            return Decimal(text)
        except InvalidOperation:
            pass

    normalized = text.replace("R$", "").replace(" ", "")
    if not MONEY_RE.match(normalized):
        return None

    try:
        return Decimal(normalized.replace(".", "").replace(",", "."))
    except InvalidOperation:
        return None


def decimal_to_float(value: object) -> float | None:
    if value is None:
        return None
    return float(value)


def convert_xls_cell(workbook, value: object, cell_type: int) -> object:
    if cell_type == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def find_header_index(norm_row: list[str], candidates: set[str]) -> int | None:
    for i, v in enumerate(norm_row):
        if v in candidates:
            return i
    return None


def get_value(row: list, index: int | None) -> object:
    if index is None or index >= len(row):
        return None
    return row[index]


def first_date_in_row(row: list) -> str | None:
    for v in row:
        parsed = parse_date_value(v)
        if parsed:
            return parsed
    return None


def extract_money_cells(row: list) -> list[Decimal]:
    non_empty_idx = [i for i, v in enumerate(row) if normalize_spaces(v) != ""]
    candidate_idx = set(non_empty_idx[-3:])
    values: list[Decimal] = []
    for i, v in enumerate(row):
        if i not in candidate_idx and not isinstance(v, str):
            continue
        parsed = parse_money_value(v)
        if parsed is not None:
            values.append(parsed)
    return values


def extract_text_parts(row: list) -> list[str]:
    parts: list[str] = []
    for v in row:
        if parse_date_value(v) or parse_money_value(v) is not None:
            continue
        text = normalize_spaces(v)
        if text:
            parts.append(text)
    return parts


def distribute_amounts(
    values: list[Decimal],
) -> tuple[Decimal | None, Decimal | None, Decimal | None]:
    if not values:
        return None, None, None
    if len(values) == 1:
        return None, None, values[0]
    if len(values) == 2:
        return values[0], None, values[1]
    return values[-3], values[-2], values[-1]


def collect_text_fragments(row: list, ignored: set[int]) -> str:
    parts: list[str] = []
    for i, v in enumerate(row):
        if i in ignored:
            continue
        if parse_money_value(v) is not None or parse_date_value(v):
            continue
        text = normalize_spaces(v)
        if text:
            parts.append(text)
    return " ".join(parts)


def append_extra_description(current: dict, row: list, ignored: set[int]) -> None:
    extra = collect_text_fragments(row, ignored)
    if extra:
        current["Descricao"] = join_description(current["Descricao"], extra)


def join_description(current: object, extra: str) -> str:
    base = normalize_spaces(current)
    if not base:
        return extra
    return f"{base} {extra}".strip()


# ── PDF-specific helpers ──────────────────────────────────────────────────────

def extract_pdf_lines(page) -> list[tuple[float, list[tuple[float, str]]]]:
    words = page.extract_words(use_text_flow=True)
    grouped: dict[float, list[tuple[float, str]]] = {}
    for word in words:
        key = round(word["top"], 1)
        grouped.setdefault(key, []).append((round(word["x0"], 1), word["text"]))
    return sorted(grouped.items(), key=lambda item: item[0])


def split_mvto_and_account(value: str) -> tuple[str, str]:
    m = re.match(r"^(\d{8})(.+)$", value)
    if m:
        return m.group(1), m.group(2)
    return value[:8], value[8:]


def clean_diario_account(value: str) -> str:
    cleaned = normalize_spaces(value)
    cleaned = re.sub(r"\s*-\s*", " - ", cleaned)
    return normalize_spaces(cleaned)


